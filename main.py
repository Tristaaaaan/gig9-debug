from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from openpyxl import load_workbook
import re


class AplicativoPersiana(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical', padding=20, spacing=15)

        self.largura_input = TextInput(hint_text="Largura", font_size=20, foreground_color=(
            1, 1, 1), background_color=(0, 0, 0), padding=10)
        self.layout.add_widget(self.largura_input)

        self.altura_input = TextInput(hint_text="Altura", font_size=20, foreground_color=(
            1, 1, 1), background_color=(0, 0, 0))
        self.layout.add_widget(self.altura_input)

        self.codigo_input = TextInput(hint_text="Qual o código do tecido:", font_size=20, foreground_color=(
            1, 1, 1), background_color=(0, 0, 0))
        self.layout.add_widget(self.codigo_input)

        self.resultado_label = Label(
            text="", font_size=20, size_hint_y=5, height=100)
        self.layout.add_widget(self.resultado_label)

        # Add a button to trigger the Excel file loading and regex search
        self.buscar_button = Button(text="Buscar", on_press=self.buscar_codigo)
        self.layout.add_widget(self.buscar_button)

        return self.layout

    def buscar_codigo(self, instance):
        # Get the code input by the user
        codigo = self.codigo_input.text

        # Load the Excel file (assuming it's in the same directory as your script)
        workbook = load_workbook('Pv.xlsx')
        sheet = workbook.active

        # Search for the code using a regular expression
        result = None
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value and re.search(codigo, str(cell_value)):
                    result = "Found: " + str(cell_value)
                    break

        # Update the result label with the search result
        if result:
            self.resultado_label.text = result
        else:
            self.resultado_label.text = "Code not found in the Excel file"


if __name__ == '__main__':
    AplicativoPersiana().run()
